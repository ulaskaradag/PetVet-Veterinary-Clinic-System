import tkinter as tk
from tkinter import messagebox
import sqlite3
from openpyxl import load_workbook
import os


# init_db is called in ClinicApp Class __init__ method for once
def init_db():
    conn = sqlite3.connect(os.path.join("db", "credentials.db"))
    cur = conn.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS users (
            user_id TEXT PRIMARY KEY,
            full_name TEXT NOT NULL,
            password TEXT,
            role TEXT NOT NULL CHECK(role IN ('veterinarian', 'secretary'))
        )
    """)

    excel_path = os.path.join("excel_files", "worker_list.xlsx")
    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
        ws = wb.active

        for user_id, full_name, role in ws.iter_rows(min_row=2, values_only=True):
            user_id = str(user_id).strip()
            full_name = str(full_name).strip()
            role = str(role).strip().lower()

            cur.execute("""
                INSERT OR IGNORE INTO users (user_id, full_name, role)
                VALUES (?, ?, ?)
            """, (user_id, full_name, role))

    conn.commit()
    conn.close()


def login_user(user_id, password):
    conn = sqlite3.connect(os.path.join("db", "credentials.db"))
    cur = conn.cursor()

    cur.execute("""
        SELECT role
        FROM users
        WHERE user_id=? AND password=?
    """, (user_id.strip(), password.strip()))

    row = cur.fetchone()
    conn.close()

    return row[0] if row else None


def register_user(user_id, password) -> bool:
    conn = sqlite3.connect(os.path.join("db", "credentials.db"))
    cur = conn.cursor()

    cur.execute("SELECT password FROM users WHERE user_id=?", (user_id,))
    row = cur.fetchone()

    if not row or row[0] is not None:
        conn.close()
        return False

    cur.execute("UPDATE users SET password=? WHERE user_id=?", (password, user_id))

    conn.commit()
    conn.close()
    return True


def reset_user_for_reregister(user_id) -> bool:
    conn = sqlite3.connect(os.path.join("db", "credentials.db"))
    cur = conn.cursor()

    cur.execute("SELECT user_id FROM users WHERE user_id=?", (user_id,))
    if not cur.fetchone():
        conn.close(); return False

    cur.execute("UPDATE users SET password=NULL WHERE user_id=?", (user_id,))
    conn.commit()
    conn.close()
    return True


class LoginFrame(tk.Frame):
    
    def __init__(self, parent, controller):
        super().__init__(parent)
        self.controller = controller

        self.title_label = tk.Label(self, font=("Arial", 18, "bold"))
        self.title_label.pack(pady=10)

        self.user_label = tk.Label(self)
        self.user_label.pack()
        self.user_entry = tk.Entry(self)
        self.user_entry.pack()

        self.pass_label = tk.Label(self)
        self.pass_label.pack()
        self.pass_entry = tk.Entry(self, show="*")
        self.pass_entry.pack()

        btn_frame = tk.Frame(self)
        btn_frame.pack(pady=20)

        self.login_btn = tk.Button(btn_frame, width=12, command=self.login)
        self.login_btn.grid(row=0, column=0, padx=5)

        self.register_btn = tk.Button(btn_frame, width=12, command=lambda: controller.show_frame("RegisterFrame"))
        self.register_btn.grid(row=0, column=1, padx=5)

        self.change_password_btn = tk.Button(self, command=self.change_password)
        self.change_password_btn.pack()

        self.update_texts()

    def update_texts(self):
        t = self.controller.language.translate
        self.title_label.config(text=t("login"))
        self.user_label.config(text=t("user_id"))
        self.pass_label.config(text=t("password"))
        self.login_btn.config(text=t("login"))
        self.register_btn.config(text=t("register"))
        self.change_password_btn.config(text=t("change_password"))

    def login(self):
        t = self.controller.language.translate

        user_id = self.user_entry.get().strip()
        password = self.pass_entry.get()

        role = login_user(user_id, password)
        if not role:
            messagebox.showerror(t("error"), t("invalid_credentials")); return

        self.controller.current_user_id = user_id
        self.controller.current_role = role

        conn = sqlite3.connect(os.path.join("db", "credentials.db"))
        cur = conn.cursor()
        cur.execute("SELECT full_name FROM users WHERE user_id=?", (user_id,))
        row = cur.fetchone()
        conn.close()

        self.controller.current_user_name = row[0] if row else user_id

        success_msg = (t("login_success_vet") if role == "veterinarian" else t("login_success_sec"))

        messagebox.showinfo(t("success"), success_msg)

        self.controller.frames["AppointmentFrame"].apply_role(role)
        self.controller.show_frame("DashboardFrame")

    def change_password(self):
        t = self.controller.language.translate

        user_id = self.user_entry.get().strip()
        if not user_id:
            messagebox.showwarning(t("warning"), t("user_id_required")); return

        conn = sqlite3.connect(os.path.join("db", "credentials.db"))
        cur = conn.cursor()
        cur.execute("SELECT user_id FROM users WHERE user_id=?", (user_id,))
        exists = cur.fetchone() is not None
        conn.close()

        if not exists:
            messagebox.showerror(t("error"), t("user_not_found")); return

        if not messagebox.askyesno(t("change_password"), t("confirm_reset")): return

        if reset_user_for_reregister(user_id):
            messagebox.showinfo(t("success"), t("reset_success"))
            self.controller.show_frame("RegisterFrame")
        else: messagebox.showerror(t("error"), t("reset_failed"))


class RegisterFrame(tk.Frame):

    def __init__(self, parent, controller):
        super().__init__(parent)
        self.controller = controller

        self.title_label = tk.Label(self, font=("Arial", 18, "bold"))
        self.title_label.pack(pady=10)

        self.user_label = tk.Label(self)
        self.user_label.pack()
        self.user_entry = tk.Entry(self)
        self.user_entry.pack()

        self.pass_label = tk.Label(self)
        self.pass_label.pack()
        self.pass_entry = tk.Entry(self, show="*")
        self.pass_entry.pack()

        btn_frame = tk.Frame(self)
        btn_frame.pack(pady=20)

        self.register_btn = tk.Button(btn_frame, width=12, command=self.register)
        self.register_btn.grid(row=0, column=0, padx=5)

        self.back_btn = tk.Button(btn_frame, width=12, command=lambda: controller.show_frame("LoginFrame"))
        self.back_btn.grid(row=0, column=1, padx=5)

        self.update_texts()

    def update_texts(self):
        t = self.controller.language.translate
        self.title_label.config(text=t("register"))
        self.user_label.config(text=t("user_id"))
        self.pass_label.config(text=t("password"))
        self.register_btn.config(text=t("register"))
        self.back_btn.config(text=t("back"))

    def register(self):
        t = self.controller.language.translate

        user_id = self.user_entry.get().strip()
        password = self.pass_entry.get()

        if not (6 <= len(password) <= 20):
            messagebox.showerror(t("error"), t("password_length_invalid")); return
            
        success = register_user(user_id, password)

        if success:
            messagebox.showinfo(t("success"), t("register_success"))
            self.controller.show_frame("LoginFrame")
        else: messagebox.showerror(t("error"), t("register_failed"))